# -*- coding: utf-8 -*-
"""One-shot extractor: Excel HSE 2026 -> TypeScript catalog (no requisitos/vencidas UI)."""
from pathlib import Path
import json
import re
import unicodedata

import openpyxl

EXCEL = Path(r"c:\Proyectos\Nov-Campus\docs\Matriz de Competencias HSE NOV EPS 2026 - WELLCHEK (1).xlsm")
OUT = Path(r"c:\Proyectos\Nov-Campus\src\lib\capacitacion\matrix-catalog.ts")

SHORT = {
    "Programa de Aire": "Programa de Aire",
    "Conciencia Ambiental": "Conciencia Ambiental",
    "Control de Documentos e Informacion de HSE": "Control de Documentos HSE",
    "Acciones Correctivas y Preventivas (CAPA)": "Acciones Correctivas y Preventivas (CAPA)",
    "Seguridad Basada en el Comportamiento (Observation Program) // HSMS.043 / Autoridad de Detener el Trabajo": "Seguridad Basada en el Comportamiento / SWA",
    "Identificacion de Peligros y Evaluacion de Riesgos": "Identificacion de Peligros y Evaluacion de Riesgos",
    "Plan de Respuesta ante Eemergencias / HSMS.010 /Primeros Axulios & RCP/DEA / HSMSWT-002 / Patogenos Trasnmitidos por la Sangre": "Plan de Respuesta ante Emergencias / RCP",
    "Reporte y Gestion de Incidentes HSE (Todos)": "Reporte y Gestion de Incidentes HSE",
    "Reporte, Gestion e Investigacion de Incidentes HSE (supervisores / investigadores (rol))": "Investigacion de Incidentes HSE (Supervisores)",
    "Gruas y Equipos de Izaje (Todos)": "Gruas y Equipos de Izaje (Todos)",
    "HSMS.007 / Gruas y Equipos de Izaje (Operadores)": "Gruas y Equipos de Izaje (Operadores)",
    "Proteccion Contra Incendios (Todos)": "Proteccion Contra Incendios (Todos)",
    "Seguridad Electrica (Todos)": "Seguridad Electrica (Todos)",
    "Camiones Industriales Motorizados (PIT) (Todos)": "Camiones Industriales (PIT)",
    "Control de Energias Peligrosas (LOTO) (Todos)": "Control de Energias Peligrosas LOTO (Todos)",
    "Control de Energias Peligrosas (LOTO) (Empleados Autorizados)": "Control de Energias Peligrosas LOTO (Autorizados)",
    "Seguridad aplicada al Pozo & Yacimiento": "Seguridad aplicada al Pozo y Yacimiento",
    "Las Reglas que Salvan Vidas": "Las Reglas que Salvan Vidas",
    "Resguardo de Maquinas & Equipos": "Resguardo de Maquinas y Equipos",
    "Conduccion y Gestion de Viajes (Todos)": "Conduccion y Gestion de Viajes (Todos)",
    "Conduccion y Gestion de Viajes (Conductores)": "Conduccion y Gestion de Viajes (Conductores)",
    "Patogenos Transmitidos por la Sangre": "Patogenos Transmitidos por la Sangre",
    "Conservacion Auditiva": "Conservacion Auditiva",
    "Ergonomia //  / Levantamiento & Seguridad en la Espalda": "Ergonomia / Levantamiento y Seguridad en la Espalda",
    "Superficies Elevadas (Todos)": "Superficies Elevadas (Todos)",
    "Seguridad Quimica (Todos)": "Seguridad Quimica (Todos)",
    "Seguridad para Trabajos en Caliente": "Seguridad para Trabajos en Caliente",
    "Seguridad con la Radiacion": "Seguridad con la Radiacion",
    "Equipo de Proteccion Personal (PPE)": "Equipo de Proteccion Personal (EPP)",
    "Seguridad con Escaleras": "Seguridad con Escaleras",
    "Seguridad con Escaleras // Estantes de Almacenamiento": "Seguridad con Escaleras y Estantes",
    "Responsabilidades del Supervisor": "Responsabilidades del Supervisor",
    "Espacios Confinados (Todos)": "Espacios Confinados (Todos)",
    "Programa de Agua (Todos)": "Programa de Agua (Todos)",
    "HIV/SIDA y otras enfermedades de transmision sexual (ETS)": "HIV/SIDA y ETS",
    "Drogas de abuso (concientizacion y prevencion)": "Drogas de abuso",
    "Vida saludable (habitos, autocuidado y bienestar)": "Vida saludable",
    "Primeros auxilios y Reanimacion Cardio Pulmonar (RCP)": "Primeros auxilios y RCP",
    "Prevencion cardiovascular (factores de riesgo y senales de alarma)": "Prevencion cardiovascular",
    "Efectos del tabaco sobre la salud": "Efectos del tabaco sobre la salud",
}


def clean(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip()


def slugify(text: str) -> str:
    nfkd = unicodedata.normalize("NFKD", text)
    ascii_text = nfkd.encode("ascii", "ignore").decode("ascii")
    ascii_text = ascii_text.lower()
    ascii_text = re.sub(r"[^a-z0-9]+", "-", ascii_text).strip("-")
    return ascii_text[:72]


def room_for(title: str) -> str:
    t = slugify(title)
    if re.search(r"ambiental|aire|agua|quimic", t):
        return "medio-ambiente"
    if re.search(r"conduccion|viaje|vial", t):
        return "seguridad-vial"
    if re.search(r"emergenc|incendio|rcp|primeros-aux|brigad", t):
        return "emergencias-respuestas"
    if re.search(r"incidente|supervisor|documento|capa|observacion|comportamiento", t):
        return "gestion-hse"
    if re.search(r"hiv|sida|droga|saludable|tabaco|cardiovascular|ergonom|auditiv|patogen", t):
        return "salud-ocupacional"
    if re.search(r"izaje|grua|loto|electric|pit|montacarga|pozo|escalera|epp|ppe|radiacion|confinado|superficie|maquina", t):
        return "seguridad-higiene"
    return "competencias-tecnicas"


def ts_str(s: str) -> str:
    return json.dumps(s, ensure_ascii=False)


wb = openpyxl.load_workbook(EXCEL, data_only=True, read_only=True, keep_vba=False)

ws = wb["FID"]
sedes = []
for row in ws.iter_rows(min_row=2, max_row=20, max_col=1, values_only=True):
    name = clean(row[0])
    if name:
        sedes.append({"code": slugify(name), "name": name})

ws = wb["LISTADO_TAREAS"]
tareas = []
for row in ws.iter_rows(min_row=2, max_row=20, max_col=1, values_only=True):
    name = clean(row[0])
    if name:
        tareas.append({"code": slugify(name), "name": name})

ws = wb["LISTADO_PERFILES PUESTO"]
puestos = []
seen = set()
for row in ws.iter_rows(min_row=2, max_row=80, max_col=3, values_only=True):
    name = clean(row[2])
    if name and name not in seen:
        seen.add(name)
        puestos.append({"code": slugify(name), "name": name})

ws = wb["Criterio de matriz competencias"]
criterios = {}
for row in ws.iter_rows(min_row=2, max_row=80, max_col=16, values_only=True):
    title = clean(row[0])
    if not title:
        continue
    vig = row[12]
    hours = row[13]
    try:
        vig_n = int(vig) if vig not in (None, "") else 0
    except Exception:
        vig_n = 0
    try:
        hours_n = int(hours) if hours not in (None, "") else 1
    except Exception:
        hours_n = 1
    criterios[slugify(title)] = {
        "title": title,
        "tipo": clean(row[11]),
        "vigencia": vig_n,
        "hours": hours_n,
        "justificacion": clean(row[1]),
        "alcance": clean(row[5]),
    }

ws = wb["Requisito x Tarea"]
rows = list(ws.iter_rows(min_row=1, max_row=80, max_col=70, values_only=True))
headers = [clean(c) for c in rows[0]]
req_pairs = []
assigned_titles = set()
for row in rows[1:]:
    sector = clean(row[0])
    tarea = clean(row[1])
    if not sector or not tarea:
        continue
    topics = []
    for idx, h in enumerate(headers[2:], start=2):
        if not h:
            continue
        val = clean(row[idx]).upper()
        if val in ("X", "1", "SI", "TRUE") or (val.startswith("S") and val != "SIN"):
            topics.append(h)
            assigned_titles.add(h)
    req_pairs.append({"sector": sector, "tarea": tarea, "topics": topics})

ws = wb["Matriz de Comp. Personas"]
people = []
for row in ws.iter_rows(min_row=8, max_row=120, max_col=13, values_only=True):
    puesto = clean(row[4])
    sector = clean(row[5])
    if not puesto or not sector:
        continue
    person_tareas = [clean(row[c]) for c in (9, 10, 11, 12) if clean(row[c])]
    people.append({"puesto": puesto, "sector": sector, "tareas": person_tareas})

wb.close()

topics = []
topic_by_norm = {}
for title in sorted(assigned_titles):
    key = slugify(title)
    crit = criterios.get(key)
    if not crit:
        # try loose match
        for ck, cv in criterios.items():
            if ck[:40] == key[:40]:
                crit = cv
                break
    vigencia = crit["vigencia"] if crit else 0
    if vigencia <= 0:
        continue
    short = SHORT.get(title) or SHORT.get(crit["title"] if crit else "") or title[:72]
    code = slugify(short)
    item = {
        "code": code,
        "title": short,
        "sourceTitle": title,
        "validityDays": vigencia,
        "hours": crit["hours"] if crit else 1,
        "roomSlug": room_for(short),
        "justificacion": (crit["justificacion"] if crit else "HSE corporativo")[:120],
    }
    topics.append(item)
    topic_by_norm[slugify(title)] = code
    topic_by_norm[key] = code

cells = []
for req in req_pairs:
    puestos_at = sorted({p["puesto"] for p in people if p["sector"] == req["sector"]})
    for puesto in puestos_at:
        has_tarea = any(
            p["puesto"] == puesto and p["sector"] == req["sector"] and req["tarea"] in p["tareas"]
            for p in people
        )
        if not has_tarea:
            continue
        codes = []
        for t in req["topics"]:
            c = topic_by_norm.get(slugify(t))
            if c and c not in codes:
                codes.append(c)
        if not codes:
            continue
        cells.append({
            "sedeCode": slugify(req["sector"]),
            "puestoCode": slugify(puesto),
            "tareaCode": slugify(req["tarea"]),
            "topicCodes": codes,
        })

lines = []
w = lines.append
w("/** Generated from the HSE 2026 WELLCHEK workbook. Do not import requisitos/vencidas sheets. */")
w("")
w("export type MatrixRoomSlug =")
w('  | "medio-ambiente"')
w('  | "salud-ocupacional"')
w('  | "seguridad-higiene"')
w('  | "seguridad-vial"')
w('  | "emergencias-respuestas"')
w('  | "gestion-hse"')
w('  | "competencias-tecnicas";')
w("")
w("export type MatrixMaster = { code: string; name: string };")
w("export type MatrixTopic = {")
w("  code: string;")
w("  title: string;")
w("  sourceTitle: string;")
w("  validityDays: number;")
w("  hours: number;")
w("  roomSlug: MatrixRoomSlug;")
w("  justificacion: string;")
w("};")
w("export type MatrixCellDef = {")
w("  sedeCode: string;")
w("  puestoCode: string;")
w("  tareaCode: string;")
w("  topicCodes: string[];")
w("};")
w("")
w(f"export const MATRIX_YEAR = 2026;")
w("")
w("export const MATRIX_SEDES: MatrixMaster[] = [")
for s in sedes:
    w(f"  {{ code: {ts_str(s['code'])}, name: {ts_str(s['name'])} }},")
w("];")
w("")
w("export const MATRIX_PUESTOS: MatrixMaster[] = [")
for s in puestos:
    w(f"  {{ code: {ts_str(s['code'])}, name: {ts_str(s['name'])} }},")
w("];")
w("")
w("export const MATRIX_TAREAS: MatrixMaster[] = [")
for s in tareas:
    w(f"  {{ code: {ts_str(s['code'])}, name: {ts_str(s['name'])} }},")
w("];")
w("")
w("export const MATRIX_TOPICS: MatrixTopic[] = [")
for t in topics:
    w("  {")
    w(f"    code: {ts_str(t['code'])},")
    w(f"    title: {ts_str(t['title'])},")
    w(f"    sourceTitle: {ts_str(t['sourceTitle'])},")
    w(f"    validityDays: {t['validityDays']},")
    w(f"    hours: {t['hours']},")
    w(f"    roomSlug: {ts_str(t['roomSlug'])},")
    w(f"    justificacion: {ts_str(t['justificacion'])},")
    w("  },")
w("];")
w("")
w("export const MATRIX_CELLS: MatrixCellDef[] = [")
for c in cells:
    codes = ", ".join(ts_str(x) for x in c["topicCodes"])
    w("  {")
    w(f"    sedeCode: {ts_str(c['sedeCode'])},")
    w(f"    puestoCode: {ts_str(c['puestoCode'])},")
    w(f"    tareaCode: {ts_str(c['tareaCode'])},")
    w(f"    topicCodes: [{codes}],")
    w("  },")
w("];")
w("")
w("export const MATRIX_DEMO_STUDENTS = [")
w('  { dni: "30111222", firstName: "Ana", lastName: "Vega", email: "ana.vega@demo.nov", sedeCode: "tb-ar-crv-wellchek", puestoCode: "supervisor", tareaCode: "supervisor" },')
w('  { dni: "30111333", firstName: "Luis", lastName: "Moreno", email: "luis.moreno@demo.nov", sedeCode: "tb-ar-crv-wellchek", puestoCode: "operador-esp-en-ensayo-no-destructivo", tareaCode: "operador-wch" },')
w('  { dni: "30111444", firstName: "Marta", lastName: "Ruiz", email: "marta.ruiz@demo.nov", sedeCode: "tb-ar-crv-wellchek", puestoCode: "operador-foster", tareaCode: "ayudante-wch" },')
w("] as const;")
w("")

OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")
print("topics", len(topics))
print("cells", len(cells))
print("puestos", len(puestos))
print("sedes", len(sedes))
print("tareas", len(tareas))
print("wrote", OUT)
